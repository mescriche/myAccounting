__author__ = 'Manuel Escriche'
import os, re, copy
import openpyxl
from datetime import datetime, date, time
from calendar import monthrange

class ExcelReader:
    def __init__(self, path, filename):
        print(self.__class__.__name__)
        basename, self._ext = os.path.splitext(filename)
        try: self._wb = openpyxl.load_workbook(os.path.join(path, filename), data_only=True)
        except: raise
        self._sheet = self._wb.worksheets[0]
                                  
    @property
    def account(self):
        return self._account
    
    @property
    def downloaded_on(self):
        return self._download_date

    @property
    def filename(self):
        return '{}_{}_{}{}'.format(self._entity_name, self._account, self._download_date.strftime('%Y%m%d'), self._ext)

    @property
    def data(self):
        return self._data

    @property
    def size(self):
        return len(self._data)
    
    def __repr__(self):
        return 'Reader({0._entity_name} | {0.journal} | {0.downloaded_on} | {0.filename} | #{0.size} records)'.format(self)

    
class OpenBankAccountReader(ExcelReader):
    def __init__(self, path, filename):
        super().__init__(path, filename)
        self._entity_name = re.sub(r'AccountReader', '', self.__class__.__name__)
        if self._sheet.cell(row=4, column=4).value is not None:
            self._account = re.sub(r'\s+', '', self._sheet.cell(row=4, column=4).value)
            match = re.match(r'\d{20}', self._account)
            if match is None: raise Exception("Missing Bank Account number")
        else: raise Exception("Missing Bank Account number")
        
        match = re.search(r'\d{2}/\d{2}/\d{4}\s\d+:\d{2}', self._sheet.cell(row=3, column=10).value)
        if match is None:
            raise InappropiateFileFormat("Missing download date")
        self._download_date = datetime.strptime(match.group(0), '%d/%m/%Y %H:%M')
        
        input_data = [[cell for cell in row if cell is not None] for row in self._sheet.iter_rows(values_only=True)]        
        input_data = [row for row in input_data if len(row) == 5]
        
        self._data = []
        for item in input_data:
            try: _date = datetime.strptime(item[0], '%d/%m/%Y').date() if isinstance(item[0],str) else item[0].date()
            except ValueError:pass
            else :
                amount  = float(re.sub(r',', '.', re.sub(r'\.', '', item[3]))) if isinstance(item[3], str) else item[3]
                balance = float(re.sub(r',', '.', re.sub(r'\.', '', item[4]))) if isinstance(item[4], str) else item[4]
                self._data.append({'odate':_date, 'vdate':_date, 'comment':str(item[2]), 'amount': amount, 'balance': balance })

class OpenBankCardReader(ExcelReader):
    def __init__(self, path, filename):
        super().__init__(path, filename)
        self._entity_name = re.sub(r'CardReader', '', self.__class__.__name__)

        if self._sheet.cell(row=2, column=1).value is not None:
            self._account = re.sub(r'\s+', '', self._sheet.cell(row=2, column=1).value)
            match = re.match(r'\d{16}', self._journal)
            if match is None: raise Exception("Missing journal number")
        else: raise Exception("Missing journal number")

        input_data = [[cell for cell in row if cell is not None] for row in self._sheet.iter_rows(values_only=True)]
        input_data = [row for row in input_data if len(row) == 5]
        _data = []
        for item in input_data:
            try: _date = datetime.strptime(item[0], '%d/%m/%Y').date()
            except ValueError: pass
            else:
                amount = float(re.sub(r',', '.', re.sub(r'\.', '', item[2]))) if isinstance(item[2], str) else float(item[2])
                _data.append({'odate':_date, 'vdate':_date, 'comment': str(item[3]),
                              'amount': amount, 'balance': None})
        self._data = [item for item in reversed(_data)]
        self._download_date = datetime.combine(self._data[0]['odate'], time(0,0,0,0))

class OpenBankCardReader2(ExcelReader):
    def __init__(self, path, filename):
        super().__init__(path, filename)
        self._entity_name = re.sub(r'CardReader2', '', self.__class__.__name__)
        if self._sheet.cell(row=4, column=4).value is not None:
            self._account = re.sub(r'\s+', '', self._sheet.cell(row=4, column=4).value)
            match = re.match(r'\d{16}', self._journal)
            if match is None: raise Exception("Missing journal number")
        else: raise Exception("Missing journal number")
        
        match = re.search(r'\d{2}/\d{2}/\d{4}\s\d+:\d{2}', self._sheet.cell(row=3, column=13).value)
        if match is None: raise Exception("Missing download date")

        self._download_date = datetime.strptime(match.group(0), '%d/%m/%Y %H:%M')

        input_data = [[cell for cell in row if cell is not None] for row in self._sheet.iter_rows(values_only=True)]
        input_data = [row for row in input_data if len(row) == 7]

        self._data = []
        for item in input_data:
            try: _date = datetime.strptime(item[0], '%d/%m/%Y').date() if isinstance(item[0],str) else item[0].date()
            except ValueError:pass
            else:
                comment = 'Credit: COMPRA EN ' + str(item[2]) + ' EN ' + str(item[4])
                self._data.append({'odate':_date, 'vdate':_date, 'comment': comment, 'amount': item[-2], 'balance': None })
        
class IngDirectReader(ExcelReader):
    def __init__(self, path, filename):
        super().__init__(path, filename)
        self._entity_name = re.sub(r'Reader', '', self.__class__.__name__)
        if  self._sheet.cell(row=2, column=4).value is not None:
            self._account = re.sub(r'\s+', '', self._sheet.cell(row=2, column=4).value)
            match = re.match(r'(ES\d{2})?\d{20}', self._journal)
            if match is None: raise Exception("Missing journal number")
        else: raise Exception("Missing journal number")
        
        match = re.search(r'\d{2}/\d{2}/\d{4}\s\d{2}:\d{2}', self._sheet.cell(row=4, column=4).value)
        
        if match is None: raise Exception("Missing download date")

        self._download_date = datetime.strptime(match.group(0),'%d/%m/%Y %H:%M')

        input_data = [[cell for cell in row if cell is not None] for row in self._sheet.iter_rows(values_only=True)]
        input_data = [row for row in input_data if len(row) == 8]
        
        self._data = []
        for item in input_data:
            try : _date = datetime.strptime(item[0], '%d/%m/%Y').date() if isinstance(item[0], str) else item[0].date()
            except ValueError: pass
            else :
                amount  = float(re.sub(r',', '.', re.sub(r'\.', '', item[-2]))) if isinstance(item[-2], str) else float(item[-2])
                balance = float(re.sub(r',', '.', re.sub(r'\.', '', item[-1]))) if isinstance(item[-1], str) else float(item[-1])
                self._data.append({'odate': _date, 'vdate': _date, 'comment': '->'.join(item[1:-3]),
                                   'amount': amount, 'balance': balance})

class BankinterAccountReader(ExcelReader):
    def __init__(self, path, filename):
        super().__init__(path, filename)
        self._entity_name =  re.sub(r'AccountReader', '', self.__class__.__name__)
        if self._sheet.cell(row=1, column=1).value is not None:
            match = re.search(r'IBAN:(ES\d{2}\d{20})',  re.sub(r'\s+', '', self._sheet.cell(row=1, column=1).value) )
            if match is None : raise Exception("Missing journal number")
            self._account = match.group(1)
        else: raise Exception("Missing journal number")
        
        input_data = [[cell for cell in row if cell is not None] for row in self._sheet.iter_rows(values_only=True)]
        input_data = [row for row in input_data if len(row) == 5]
        _data = []
        for item in input_data:
            try:
                _datec =  datetime.strptime(item[0], '%d/%m/%Y').date() if isinstance(item[0],str) else item[0].date()
                _datev =  datetime.strptime(item[1], '%d/%m/%Y').date() if isinstance(item[1],str) else item[1].date()
            except ValueError: pass
            else:
                amount  = float(re.sub(r',', '.', re.sub(r'\.', '', item[3]))) if isinstance(item[3], str) else item[3]
                balance = float(re.sub(r',', '.', re.sub(r'\.', '', item[4]))) if isinstance(item[4], str) else item[4]
                _data.append({'odate':_datec, 'vdate':_datev, 'comment':str(item[2]), 'amount': amount, 'balance': balance })
        self._data = [item for item in reversed(_data)]
        #for item in self._data: print(item)
        self._download_date = datetime.combine(self._data[0]['odate'], time(0,0,0,0))

class BankinterCreditCardReader(ExcelReader):
    def __init__(self, path, filename):
        super().__init__(path, filename)
        self._entity_name =  re.sub(r'CreditCardReader', '', self.__class__.__name__)

        if self._sheet.cell(row=1, column=1).value is not None:
            match = re.search(r'.*:.*(\d{4})', re.sub(r'\s+', '', self._sheet.cell(row=1, column=1).value))
            if match is None: raise Exception("Missing journal number")
            self._account = match.group(1)
        else: raise Exception("Missing journal number")
                
        _input = [[cell for cell in row if cell is not None] for row in self._sheet.iter_rows(values_only=True)]
        total_credit = round(next(filter(lambda x: len(x) > 0 and x[0] == 'Total Crédito', _input))[1],2)
        
        input_data = filter(lambda x: len(x) == 4 and isinstance(x[0], datetime) , _input)
        credit_entries = list(filter(lambda x: x[2] == 'Crédito', input_data))
        credit = round(sum(map(lambda x:x[3], credit_entries)),2)        

        if credit != total_credit: raise Exception("Total sum of credit entries  don't match Total Crédito")
        _data = []
        _vdate = credit_entries[-1][0].date()
        for item in credit_entries:
            _data.append({'odate':item[0].date(), 'vdate':_vdate, 'comment':str(item[1]), 'amount':item[3], 'balance':None})

        self._data = [item for item in reversed(_data)]
        #for item in self._data: print(item)
        self._download_date = datetime.combine(self._data[0]['odate'], time(0,0,0,0))

        
def create_excel_reader(filename:str) -> ExcelReader:
    assert os.path.exists(filename), filename + " doesn't exists"
    assert filename.endswith(".xlsx"), filename + " hasn't got .xlsx extension"
    path = os.path.dirname(filename)
    fname = os.path.basename(filename)
    readers = ('OpenBankAccountReader', 'OpenBankCardReader','OpenBankCardReader2'
               'IngDirectReader',
               'BankinterAccountReader','BankinterCreditCardReader' )
    for readername in readers:
        action = f"{readername}('{path}','{fname}')"
        try: reader = eval(action)
        except Exception: pass
        else:
            print('#{} records en {} downloaded on {}'.format(len(reader.data),filename, reader.downloaded_on))
            return reader
    else:
        raise Exception(f'OpenBank, Bankinter and IngDirect excel readers failed on {filename}')

    



            
            
            

        
    
